import openpyxl

file_path = r's:\elasticcost\vm_spec_cloud_part\XpressAzure Price List v2.7 - TND.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

for sheet_name in wb.sheetnames:
    print(f"\n=== Sheet: {sheet_name} ===")
    sheet = wb[sheet_name]
    
    # Print non-empty rows in the entire sheet
    for r in range(1, sheet.max_row + 1):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 15)]
        if not any(row_vals):
            continue
        print(f"Row {r:02d}: {row_vals[:6]}")
