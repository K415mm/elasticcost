import openpyxl
import json

file_path = r's:\elasticcost\vm_spec_cloud_part\XpressAzure Price List v2.7 - TND.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

data = {}

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    
    vms = []
    disks = []
    
    # Mode variables to track what we're reading
    # We can just iterate through rows and parse
    for r in range(1, sheet.max_row + 1):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 10)]
        if not any(row_vals):
            continue
            
        # Check if row is VM header
        if row_vals[1] == 'VM' and row_vals[2] == 'vCPU':
            continue
        
        # Check if VM row
        # A VM row has a name in col 1, vCPU in col 2, RAM in col 3, Price Windows in col 4, Price Linux in col 5
        if r >= 4 and r <= 36 and row_vals[1] is not None and isinstance(row_vals[2], (int, float)):
            vms.append({
                'name': str(row_vals[1]).strip(),
                'vcpu': int(row_vals[2]),
                'ram_gb': float(row_vals[3]),
                'price_windows_tnd': float(row_vals[4]) if row_vals[4] is not None else 0.0,
                'price_linux_tnd': float(row_vals[5]) if row_vals[5] is not None else 0.0
            })
            
        # Check if Disk row
        if r >= 39 and r <= 46 and row_vals[1] is not None and isinstance(row_vals[2], (int, float)):
            disks.append({
                'name': str(row_vals[1]).strip(),
                'size_gb': float(row_vals[2]),
                'type': str(row_vals[3]).strip(),
                'price_tnd': float(row_vals[4]) if row_vals[4] is not None else 0.0
            })
            
    data[sheet_name] = {
        'vms': vms,
        'disks': disks
    }

# Save to public/assets/json/xpress_azure_prices.json
import os
os.makedirs(r's:\elasticcost\public\assets\json', exist_ok=True)
with open(r's:\elasticcost\public\assets\json\xpress_azure_prices.json', 'w') as f:
    json.dump(data, f, indent=4)

print("Parsed successfully!")
print(f"Dataxion VMs: {len(data['Dataxion']['vms'])}, Disks: {len(data['Dataxion']['disks'])}")
print(f"TT VMs: {len(data['TT']['vms'])}, Disks: {len(data['TT']['disks'])}")
